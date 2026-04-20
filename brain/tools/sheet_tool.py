"""
Sheet tool — read and write CSV / Excel files.

Handles:
  - CSV (always — stdlib only)
  - Excel (.xlsx) via openpyxl if installed
"""

from __future__ import annotations

import csv
import os
import sys
from typing import Any, List, Optional

from brain.tools import register_tool


def _app_root() -> str:
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


def _resolve_path(path: str) -> str:
    """Resolve relative paths against brain_state/sheets/."""
    if os.path.isabs(path):
        return path
    base = os.path.join(_app_root(), "brain_state", "sheets")
    os.makedirs(base, exist_ok=True)
    return os.path.join(base, path)


@register_tool(
    name="sheet.read",
    description="Read rows from a CSV or XLSX file.",
    params={
        "path": "str — file path (absolute, or relative to brain_state/sheets/)",
        "sheet": "str — for xlsx, sheet name (default: first sheet)",
        "limit": "int — max rows to return (default: 1000)",
    },
    required=["path"],
    category="data",
)
def sheet_read(path: str, sheet: Optional[str] = None, limit: int = 1000) -> dict:
    full = _resolve_path(path)
    if not os.path.exists(full):
        return {"ok": False, "error": f"File not found: {full}"}

    ext = os.path.splitext(full)[1].lower()

    try:
        if ext == ".csv":
            rows: List[dict] = []
            with open(full, "r", encoding="utf-8", newline="") as f:
                reader = csv.DictReader(f)
                for i, row in enumerate(reader):
                    if i >= limit:
                        break
                    rows.append(row)
            return {
                "ok": True, "rows": rows, "row_count": len(rows),
                "path": full, "format": "csv",
            }
        elif ext in (".xlsx", ".xlsm"):
            try:
                import openpyxl  # type: ignore
            except ImportError:
                return {
                    "ok": False,
                    "error": "openpyxl required for xlsx files. Run: pip install openpyxl",
                }
            wb = openpyxl.load_workbook(full, data_only=True, read_only=True)
            ws = wb[sheet] if sheet else wb.active
            rows_iter = ws.iter_rows(values_only=True)
            headers = next(rows_iter, None)
            if headers is None:
                return {"ok": True, "rows": [], "row_count": 0, "path": full}
            headers = [str(h) if h is not None else "" for h in headers]
            rows: List[dict] = []
            for i, r in enumerate(rows_iter):
                if i >= limit:
                    break
                rows.append({headers[j]: r[j] for j in range(len(headers)) if j < len(r)})
            wb.close()
            return {
                "ok": True, "rows": rows, "row_count": len(rows),
                "path": full, "format": "xlsx", "sheet": ws.title,
            }
        else:
            return {"ok": False, "error": f"Unsupported format: {ext}"}
    except Exception as e:
        return {"ok": False, "error": f"{type(e).__name__}: {e}"}


@register_tool(
    name="sheet.write",
    description="Write rows to a CSV or XLSX file.",
    params={
        "path": "str — file path",
        "rows": "list[dict] — rows as list of dicts, keys become column headers",
        "append": "bool — append instead of overwrite (CSV only, default: false)",
    },
    required=["path", "rows"],
    category="data",
)
def sheet_write(path: str, rows: List[dict], append: bool = False) -> dict:
    full = _resolve_path(path)
    os.makedirs(os.path.dirname(full) or ".", exist_ok=True)
    ext = os.path.splitext(full)[1].lower()

    if not rows:
        return {"ok": False, "error": "No rows provided"}
    if not isinstance(rows[0], dict):
        return {"ok": False, "error": "Each row must be a dict"}

    try:
        if ext == ".csv":
            mode = "a" if append and os.path.exists(full) else "w"
            headers = list(rows[0].keys())
            with open(full, mode, encoding="utf-8", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                if mode == "w":
                    writer.writeheader()
                writer.writerows(rows)
            return {"ok": True, "path": full, "written": len(rows)}
        elif ext == ".xlsx":
            try:
                import openpyxl  # type: ignore
            except ImportError:
                return {"ok": False, "error": "openpyxl required. pip install openpyxl"}
            if os.path.exists(full) and append:
                wb = openpyxl.load_workbook(full)
                ws = wb.active
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(list(rows[0].keys()))
            for r in rows:
                ws.append(list(r.values()))
            wb.save(full)
            return {"ok": True, "path": full, "written": len(rows)}
        else:
            return {"ok": False, "error": f"Unsupported format: {ext}"}
    except Exception as e:
        return {"ok": False, "error": f"{type(e).__name__}: {e}"}


@register_tool(
    name="sheet.list",
    description="List all sheet files in brain_state/sheets/.",
    params={},
    required=[],
    category="data",
)
def sheet_list() -> dict:
    base = os.path.join(_app_root(), "brain_state", "sheets")
    if not os.path.isdir(base):
        return {"ok": True, "files": []}
    files = []
    for f in sorted(os.listdir(base)):
        p = os.path.join(base, f)
        if os.path.isfile(p):
            files.append({
                "name": f,
                "size_bytes": os.path.getsize(p),
                "modified": os.path.getmtime(p),
            })
    return {"ok": True, "files": files, "count": len(files)}
